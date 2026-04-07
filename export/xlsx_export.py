"""XLSX Export — Raw Extraction tab.

Produces an Excel workbook with a "Raw Extraction" sheet and a "Summary" sheet.

Ported from: financial-spreadx/lib/export/xlsx-export.ts (Raw Extraction tab only)
"""

from __future__ import annotations

import json
from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

if TYPE_CHECKING:
    from pipeline.orchestrator import PipelineResult

# ── Styling (matching the TS ExcelJS styling) ────────────────────────────────

HEADER_FILL = PatternFill(start_color="1A1917", end_color="1A1917", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=9)

RAW_COLUMNS = [
    ("Page", 8),
    ("Statement Type", 18),
    ("Raw Label", 44),
    ("Indentation", 11),
    ("Is Subtotal", 11),
    ("Note Ref", 12),
    ("Scope", 14),
    ("Values (JSON)", 40),
]


def build_raw_extraction_xlsx(pipeline_result: PipelineResult) -> bytes:
    """Build an Excel workbook with Raw Extraction + Summary tabs.

    Args:
        pipeline_result: Output from run_pipeline().

    Returns:
        Excel file as bytes.
    """
    wb = Workbook()

    # ── Tab 1: Summary ───────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 42

    # Header row
    for col_idx, header in enumerate(["Field", "Value"], 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    s = pipeline_result.summary
    summary_data = [
        ("Template Type", s.get("template_type", "—")),
        ("Total Pages", s.get("total_pages", 0)),
        ("Digital Pages", s.get("digital_pages", 0)),
        ("Scanned Pages", s.get("scanned_pages", 0)),
        ("Hybrid Pages", s.get("hybrid_pages", 0)),
        ("Filtered Pages", s.get("filtered_pages", 0)),
        ("OCR Pages", s.get("ocr_pages", 0)),
        ("", ""),
        ("Total Extracted Rows", s.get("total_rows", 0)),
        ("Income Statement Rows", s.get("rows_by_type", {}).get("income_statement", 0)),
        ("Balance Sheet Rows", s.get("rows_by_type", {}).get("balance_sheet", 0)),
        ("Cash Flow Rows", s.get("rows_by_type", {}).get("cash_flow", 0)),
        ("Equity Statement Rows", s.get("rows_by_type", {}).get("equity_statement", 0)),
        ("", ""),
        ("Total Notes Extracted", s.get("total_notes", 0)),
    ]

    for row_idx, (field, value) in enumerate(summary_data, 2):
        cell_a = ws_summary.cell(row=row_idx, column=1, value=field)
        cell_b = ws_summary.cell(row=row_idx, column=2, value=value)
        cell_a.font = BODY_FONT
        cell_b.font = BODY_FONT

    # ── Tab 2: Raw Extraction ────────────────────────────────────────────
    ws_raw = wb.create_sheet("Raw Extraction")

    # Set column widths and headers
    for col_idx, (header, width) in enumerate(RAW_COLUMNS, 1):
        ws_raw.column_dimensions[chr(64 + col_idx)].width = width
        cell = ws_raw.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    # Data rows
    for row_idx, row in enumerate(pipeline_result.extracted_rows, 2):
        values = [
            row.get("page", ""),
            row.get("statement_type", ""),
            row.get("raw_label", ""),
            row.get("indentation_level", 0),
            "Yes" if row.get("is_subtotal") else "No",
            row.get("note_ref") or "",
            row.get("statement_scope", ""),
            json.dumps(row.get("raw_values", {})),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_raw.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT

    # Serialize to bytes
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
