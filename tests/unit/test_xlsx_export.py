"""Unit tests for XLSX export — Raw Extraction tab."""

from __future__ import annotations

import json
from io import BytesIO

from openpyxl import load_workbook

from export.xlsx_export import build_raw_extraction_xlsx
from pipeline.orchestrator import PipelineResult


def _make_result(rows: list[dict] | None = None) -> PipelineResult:
    """Build a minimal PipelineResult for testing."""
    if rows is None:
        rows = [
            {
                "page": 1,
                "statement_type": "income_statement",
                "raw_label": "Revenue",
                "indentation_level": 0,
                "is_subtotal": False,
                "note_ref": "Note 5",
                "statement_scope": "consolidated",
                "raw_values": {"2024": 100.5, "2023": 90.0},
            },
            {
                "page": 1,
                "statement_type": "income_statement",
                "raw_label": "Cost of Sales",
                "indentation_level": 1,
                "is_subtotal": False,
                "note_ref": None,
                "statement_scope": "consolidated",
                "raw_values": {"2024": -50.0, "2023": -45.0},
            },
            {
                "page": 2,
                "statement_type": "balance_sheet",
                "raw_label": "Total Assets",
                "indentation_level": 0,
                "is_subtotal": True,
                "note_ref": None,
                "statement_scope": "consolidated",
                "raw_values": {"2024": 5000.0},
            },
        ]
    result = PipelineResult(extracted_rows=rows)
    result.summary = {
        "total_pages": 2,
        "digital_pages": 2,
        "scanned_pages": 0,
        "hybrid_pages": 0,
        "filtered_pages": 2,
        "total_rows": len(rows),
        "rows_by_type": {"income_statement": 2, "balance_sheet": 1},
        "total_notes": 0,
        "ocr_pages": 0,
        "template_type": "T1",
    }
    return result


def _load(xlsx_bytes: bytes):
    return load_workbook(BytesIO(xlsx_bytes))


def test_sheet_name():
    wb = _load(build_raw_extraction_xlsx(_make_result()))
    assert "Raw Extraction" in wb.sheetnames


def test_header_columns():
    wb = _load(build_raw_extraction_xlsx(_make_result()))
    ws = wb["Raw Extraction"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
    assert headers == [
        "Page", "Statement Type", "Raw Label", "Indentation",
        "Is Subtotal", "Note Ref", "Scope", "Values (JSON)",
    ]


def test_row_count():
    wb = _load(build_raw_extraction_xlsx(_make_result()))
    ws = wb["Raw Extraction"]
    # 1 header + 3 data rows = max_row 4
    assert ws.max_row == 4


def test_values_json_valid():
    wb = _load(build_raw_extraction_xlsx(_make_result()))
    ws = wb["Raw Extraction"]
    # Row 2, column 8 = Values (JSON) for first data row
    val = ws.cell(row=2, column=8).value
    parsed = json.loads(val)
    assert parsed["2024"] == 100.5
    assert parsed["2023"] == 90.0


def test_empty_result():
    wb = _load(build_raw_extraction_xlsx(_make_result(rows=[])))
    ws = wb["Raw Extraction"]
    # Header only, no data rows
    assert ws.max_row == 1
