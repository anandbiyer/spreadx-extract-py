"""Integration tests — full pipeline end-to-end (requires ANTHROPIC_API_KEY)."""

from __future__ import annotations

import json
import os
from io import BytesIO

import pytest
from openpyxl import load_workbook

from export.xlsx_export import build_raw_extraction_xlsx
from pipeline.orchestrator import run_pipeline

pytestmark = [
    pytest.mark.integration,
    pytest.mark.slow,
    pytest.mark.skipif(
        not os.environ.get("ANTHROPIC_API_KEY"),
        reason="ANTHROPIC_API_KEY not set",
    ),
]


@pytest.mark.timeout(600)
def test_pipeline_lt_finance(pdf_lt_finance: bytes):
    """Full pipeline on LT Finance (T3) -> >=10 rows, multiple statement types."""
    result = run_pipeline(pdf_lt_finance, "T3")

    assert result.summary["total_rows"] >= 10
    types_found = {r["statement_type"] for r in result.extracted_rows}
    assert "income_statement" in types_found
    assert "balance_sheet" in types_found
    assert result.template_type == "T3"


@pytest.mark.timeout(600)
def test_pipeline_cash_america(pdf_cash_america: bytes):
    """Full pipeline on Cash America (T1) -> >=5 rows."""
    result = run_pipeline(pdf_cash_america, "T1")

    assert result.summary["total_rows"] >= 5


@pytest.mark.timeout(600)
def test_pipeline_xlsx_output(pdf_lt_finance: bytes):
    """Pipeline + XLSX export -> valid file, correct structure."""
    result = run_pipeline(pdf_lt_finance, "T3")
    xlsx_bytes = build_raw_extraction_xlsx(result)

    assert len(xlsx_bytes) > 0

    wb = load_workbook(BytesIO(xlsx_bytes))
    assert "Raw Extraction" in wb.sheetnames
    assert "Summary" in wb.sheetnames

    ws = wb["Raw Extraction"]
    # Header row + at least some data rows
    assert ws.max_row >= 2

    # Verify column structure
    headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
    assert headers[0] == "Page"
    assert headers[2] == "Raw Label"
    assert headers[7] == "Values (JSON)"

    # Verify a data cell has valid JSON
    val = ws.cell(row=2, column=8).value
    assert val is not None
    parsed = json.loads(val)
    assert isinstance(parsed, dict)
